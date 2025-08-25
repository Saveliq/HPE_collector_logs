import os

import json
from pathlib import Path
from openpyxl import Workbook


# === CONFIGURATION ===
OUTPUT_PATH = Path("Аудит_результаты.xlsx")
SERVER_NUMBER = 1

# Порядок и подписи полей, которые нужно извлекать из JSON
FIELDS = [
    ("SKU", "SKU"),
    ("SerialNumber", "Серийный номер"),
    ("Model", "Модель"),
    ("FirmwareVersion", "Версия прошивки"),
    ("BiosVersion", "BIOS версия"),
    ("PowerState", "Состояние питания"),
    ("ServerHealth", "Здоровье сервера"),
    ("MemorySummaryHealth", "Состояние памяти"),
    # Добавьте другие поля по мере необходимости
]

def write_proc_info(ws, server_data, start_row, start_col):
    ws.cell(row=start_row, column=start_col, value="Процессор")
    ws.cell(row=start_row, column=start_col+1, value=server_data.get("ProcessorModel"))
    ws.cell(row=start_row, column=start_col+2, value=int(server_data.get("ProcessorCount")))
    return (start_row+1, 5)

def write_NIC_info(ws, server_data, start_row, start_col):
    def perf_json(server_data):
        NICs = server_data.get("NetworkAdapters")
        # result = {"647586-B21": {"Name": "HP FlexFabric 10Gb 2-port 554FLB Adapter", "count": 1}}
        result = {}
        for NIC in NICs:
            if NIC.get("PartNumber") in result.keys():
                result[NIC.get("PartNumber")]["count"] += 1
            else:
                result[NIC.get("PartNumber")] = {"Name": NIC.get("Name"), "count": 1}

        print(result)
        return result

    write_data = perf_json(server_data)
    for PartNumber in write_data:
        ws.cell(row=start_row, column=start_col, value=PartNumber)
        ws.cell(row=start_row, column=start_col+1, value=write_data.get(PartNumber).get("Name"))
        ws.cell(row=start_row, column=start_col+2, value=write_data.get(PartNumber).get("count"))
        start_row += 1
    return (start_row, 5)

def write_RAID_info(ws, server_data, start_row, start_col):
    def perf_json(server_data):
        RAIDs = server_data.get("SmartStorage")
        spareBattery = server_data.get("StorageBattery", None)
        # result = {"Smart Array P220i Controller": {"Cache": "CacheMemorySizeMiB", "Battery": 1, "count": 1}}
        result = {}
        for RAID in RAIDs:
            if not RAID.get("LocationFormat") in ["PCISlot"]:
                continue
            if RAID.get("Model") in result.keys():
                result[RAID.get("Model")]["RAIDCount"] += 1
                if RAID.get("BackupPowerSourceStatus") == "Present":
                    result[RAID.get("Model")]["BAtteryCount"] += 1
            else:
                result[RAID.get("Model")] = {"Cache": RAID.get("CacheMemorySizeMiB"),"PartNumber": RAID.get("PartNumber"),"Battery": 1 if RAID.get("BackupPowerSourceStatus") else 0,"BAtteryCount":0,  "RAIDCount": 1, "description": str({"CacheMemorySizeMiB":RAID.get("CacheMemorySizeMiB"),"Model":RAID.get("Model"),"BackupPowerSourceStatus":RAID.get("BackupPowerSourceStatus"),"LocationFormat":RAID.get("LocationFormat")})}
                if RAID.get("BackupPowerSourceStatus") == "Present":
                    result[RAID.get("Model")]["BAtteryCount"] += 1
        print(result)
        return result, spareBattery

    write_data, spareBattery = perf_json(server_data)
    for el in write_data:
        ws.cell(row=start_row, column=start_col, value=write_data.get(el).get("PartNumber"))
        ws.cell(row=start_row, column=start_col+1, value=el)
        ws.cell(row=start_row, column=start_col+2, value=write_data.get(el).get("RAIDCount"))
        ws.cell(row=start_row, column=start_col+3, value=write_data.get(el).get("description"))
        if write_data.get(el).get("Battery"):
            start_row += 1
            ws.cell(row=start_row, column=start_col, value=spareBattery)
            ws.cell(row=start_row, column=start_col+1, value="Battery")
            ws.cell(row=start_row, column=start_col+2, value=write_data.get(el).get("BAtteryCount"))
        start_row += 1
    return (start_row, 5)


def write_MEM_info(ws, server_data, start_row, start_col):
    def perf_json(server_data):
        DIMMs = server_data.get("Memory")
        result = {}
        for DIMM in DIMMs:
            if DIMM.get("DIMMStatus") in ["NotPresent"]:
                continue
            if DIMM.get("PartNumber") in result.keys():
                result[DIMM.get("PartNumber")]["count"] += 1
            else:
                result[DIMM.get("PartNumber")] = {"SizeMB":DIMM.get("SizeMB"),"MGz":DIMM.get("Frequency"),"Rank":DIMM.get("Rank"),"Technology":DIMM.get("Technology"),"count": 1, "other": str(DIMM)}
        print(result)
        return result

    write_data = perf_json(server_data)
    for el in write_data:
        ws.cell(row=start_row, column=start_col, value=el)
        ws.cell(row=start_row, column=start_col+1, value="Память")
        ws.cell(row=start_row, column=start_col+2, value=write_data.get(el).get("count"))
        ws.cell(row=start_row, column=start_col+3, value=write_data.get(el).get("other"))
        start_row += 1
    return (start_row, 5)

def write_PSU_info(ws, server_data, start_row, start_col):
    def perf_json(server_data):
        PSUs = server_data.get("PowerSupplies")
        result = {}
        for PSU in PSUs:
            if PSU.get("SparePartNumber") in result.keys():
                result[PSU.get("SparePartNumber")]["count"] += 1
            else:
                result[PSU.get("SparePartNumber")] = {"count": 1, "other": str(PSU)}
        return result

    write_data = perf_json(server_data)
    for el in write_data:
        ws.cell(row=start_row, column=start_col, value=el)
        ws.cell(row=start_row, column=start_col+1, value="Блок питания")
        ws.cell(row=start_row, column=start_col+2, value=write_data.get(el).get("count"))
        ws.cell(row=start_row, column=start_col+3, value=write_data.get(el).get("other"))
        start_row += 1
    return (start_row, 5)

def write_DISK_info(ws, server_data, start_row, start_col):
    def perf_json(server_data):
        DISKs = server_data.get("PhysicalDisks")
        result = {}
        for DISK in DISKs:
            if DISK.get("Model") in result.keys():
                result[DISK.get("Model")]["count"] += 1
            else:
                result[DISK.get("Model")] = {'CapacityGB': DISK.get("CapacityGB"),
                                             'InterfaceSpeedMbps': DISK.get("InterfaceSpeedMbps"),
                                             'InterfaceType': DISK.get("InterfaceType"),
                                             'MediaType': DISK.get("MediaType"), 'count': 1, "other": str(DISK)}
        print(result)
        return result

    write_data = perf_json(server_data)
    for el in write_data:
        ws.cell(row=start_row, column=start_col, value=el)
        ws.cell(row=start_row, column=start_col+1, value=write_data.get(el).get("MediaType"))
        ws.cell(row=start_row, column=start_col+2, value=write_data.get(el).get("count"))
        ws.cell(row=start_row, column=start_col+3, value=write_data.get(el).get("other"))
        start_row += 1
    return (start_row, 5)

def write_other_info(ws, server_data, start_row, start_col):
    if not server_data.get("SDCard", "") == "Absent" :
        ws.cell(row=start_row, column=start_col, value="None")
        ws.cell(row=start_row, column=start_col + 1, value="SDCard")
        ws.cell(row=start_row, column=start_col + 2, value=1)
        start_row += 1
    if not server_data.get("TrustedModules", "") in ["NotPresent", "Absent"] :
        ws.cell(row=start_row, column=start_col, value="None")
        ws.cell(row=start_row, column=start_col + 1, value="TrustedModule")
        ws.cell(row=start_row, column=start_col + 2, value=1)
        start_row += 1
    return (start_row, 5)

def extract_json_from_text(text: str) -> dict:
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start = text.find('{')
        end = text.rfind('}') + 1
        try:
            return json.loads(text[start:end])
        except Exception:
            return {}

def write_desc_info(ws, start_row, current_col):
    Values = ["№", "S/N", "P/N", "Наименование", "Spare P/N", "Description", "Quantity", "Комментарий"]
    for column in range(len(Values)):
        ws.cell(row=start_row, column=current_col + column, value=Values[column])
    return (start_row + 1, 1)


def write_server_info(ws, server_data, start_row, current_col):
    ws.cell(row=start_row, column=1, value=SERVER_NUMBER)
    ws.cell(row=start_row, column=2, value=server_data.get("SerialNumber"))
    ws.cell(row=start_row, column=3, value=server_data.get("SKU"))
    ws.cell(row=start_row, column=4, value=server_data.get("Model"))
    start_row += 1
    return (start_row, 4+1)
from getFromJSON.getArrayControllers import get_array_controllers
from getFromJSON.getChassis import get_chassis
from getFromJSON.getEmbeddedMedia import get_embedded_media
from getFromJSON.getManager import get_manager
from getFromJSON.getMemory import get_memory
from getFromJSON.getNetworkAdapters import get_network_adapters
from getFromJSON.getPCISlots import get_PCI_slots
from getFromJSON.getPower import get_power
from getFromJSON.getSystem import get_system
import json
def print_parser(file_name):
    print("-" * 50)
    with open(file_name, 'r', encoding='utf-8') as f:

        try:
            raw_data = json.load(f)
            if not raw_data.get('/redfish/v1/Chassis/1', None):
                return None
        except Exception as e:
            return None
        data = None
        data = get_chassis(raw_data)
        data.update(get_manager(raw_data))
        data.update(get_system(raw_data))
        data["PowerSupplies"] = get_power(raw_data)
        data.update(get_embedded_media(raw_data))
        data["NetworkAdapters"] = get_network_adapters(raw_data)
        data.update(get_PCI_slots(raw_data))
        data["SmartStorage"], data["PhysicalDisks"] = get_array_controllers(raw_data)
        data.update(get_power(raw_data))
        data["Memory"] = get_memory(raw_data)
        print(json.dumps(data, indent=4))
        return data
    print("-"*50)

def _parseJSONToExel(json_files, folder_selected):
    global SERVER_NUMBER
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Аудит"
    current_row = 1
    current_col = 1

    current_row, current_col = write_desc_info(ws_out, current_row, 1)
    for json_file in json_files:
        print(f"Обрабатывается файл: {json_file}")
        data = print_parser(str(json_file))
        if data is None:
            continue
        # запись информации о сервере
        current_row, current_col = write_server_info(ws_out, data, current_row, current_col)
        current_row, current_col = write_proc_info(ws_out, data, current_row, current_col)
        current_row, current_col = write_NIC_info(ws_out, data, current_row, current_col)
        current_row, current_col = write_RAID_info(ws_out, data, current_row, current_col)
        current_row, current_col = write_MEM_info(ws_out, data, current_row, current_col)
        current_row, current_col = write_DISK_info(ws_out, data, current_row, current_col)
        current_row, current_col = write_PSU_info(ws_out, data, current_row, current_col)
        current_row, current_col = write_other_info(ws_out, data, current_row, current_col)

        # пустая строка между серверами
        current_row += 2
        SERVER_NUMBER += 1
    wb_out.save(os.path.join(folder_selected, OUTPUT_PATH))
    # wb_out.save(folder_selecteded+OUTPUT_PATH)
    print(f"✅ Аудит сохранён: {OUTPUT_PATH}")


